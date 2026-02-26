import streamlit as st
import pandas as pd
import os
import io
import re
import openpyxl
from datetime import datetime

# 페이지 기본 설정
st.set_page_config(page_title="알러지 양식 변환기", layout="wide")

# ==========================================
# 1. 변환 로직 함수 정의
# ==========================================

def extract_cas(text):
    """텍스트 내에서 다른 데이터나 안내문구와 혼동되지 않도록 CAS NO 형식만 정확히 추출합니다."""
    if pd.isna(text):
        return []
    # CAS NO 정규식: 숫자2~7자리-숫자2자리-숫자1자리
    return re.findall(r'\b\d{2,7}-\d{2}-\d\b', str(text))

def logic_cff_83(input_df, template_path, customer_name, product_name):
    """CFF 모드 -> 83 CFF 변환 로직"""
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # 1. 양식 C열의 수식들부터 모두 제거
    for row in ws.iter_rows(min_col=3, max_col=3, min_row=1):
        for cell in row:
            if str(cell.value).startswith('='):
                cell.value = None

    # 2. "Sheet2" 시트 삭제 (순서 무조건 준수)
    if "Sheet2" in wb.sheetnames:
        del wb["Sheet2"]

    # 3. 원본(F열)과 양식(B열) CAS NO 대조
    source_data = {}
    # 원본 데이터 순회 (F열 인덱스: 5, L열 인덱스: 11)
    for idx, row in input_df.iterrows():
        cas_text = row.iloc[5] if len(row) > 5 else None
        val = row.iloc[11] if len(row) > 11 else None
        
        cas_list = extract_cas(cas_text)
        for cas in cas_list:
            source_data[cas] = val

    # 양식 C열에 복사
    for r in range(1, ws.max_row + 1):
        template_cas_text = ws.cell(row=r, column=2).value
        if template_cas_text:
            template_cas_list = extract_cas(template_cas_text)
            for t_cas in template_cas_list:
                # 한 셀의 여러 CAS NO 중 하나라도 일치하면 동일 물질로 인식
                if t_cas in source_data:
                    ws.cell(row=r, column=3).value = source_data[t_cas]
                    break 

    # 4. 고객사명, 제품명, 현재 날짜 입력
    ws['B9'] = customer_name
    ws['B10'] = product_name
    ws['E10'] = datetime.now().strftime("%Y-%m-%d")

    return wb

def logic_cff_26(input_df, template_path, customer_name, product_name):
    """CFF 모드 -> 26 통합 변환 로직"""
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # 원본(F열)과 양식(B열) CAS NO 대조
    source_data = {}
    for idx, row in input_df.iterrows():
        cas_text = row.iloc[5] if len(row) > 5 else None
        val = row.iloc[11] if len(row) > 11 else None
        
        cas_list = extract_cas(cas_text)
        for cas in cas_list:
            source_data[cas] = val

    # 양식 C열에 복사
    for r in range(1, ws.max_row + 1):
        template_cas_text = ws.cell(row=r, column=2).value
        if template_cas_text:
            template_cas_list = extract_cas(template_cas_text)
            for t_cas in template_cas_list:
                if t_cas in source_data:
                    ws.cell(row=r, column=3).value = source_data[t_cas]
                    break

    # 고객사명, 제품명, 현재 날짜 입력
    ws['B11'] = customer_name
    ws['B12'] = product_name
    ws['E13'] = datetime.now().strftime("%Y-%m-%d")

    return wb

def logic_hp_83(input_df, template_path, customer_name, product_name):
    """HP 모드 -> 83 HP 변환 로직"""
    # TODO: 차후 구현될 HP 로직을 위해 파라미터만 맞춰둠
    return openpyxl.load_workbook(template_path)

def logic_hp_26(input_df, template_path, customer_name, product_name):
    """HP 모드 -> 26 통합 변환 로직"""
    # TODO: 차후 구현될 HP 로직을 위해 파라미터만 맞춰둠
    return openpyxl.load_workbook(template_path)

# 엑셀 다운로드를 위한 바이너리 변환 함수 (openpyxl 객체 호환 추가)
def to_excel(data):
    output = io.BytesIO()
    if isinstance(data, pd.DataFrame):
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            data.to_excel(writer, index=False, sheet_name='Sheet1')
    else:
        # 양식 파일(openpyxl workbook)인 경우 그대로 저장
        data.save(output)
    processed_data = output.getvalue()
    return processed_data

# ==========================================
# 2. UI 레이아웃 구성
# ==========================================

st.title("📄 알러지 양식 변환기")
st.markdown("---")

# [상단] 입력 및 설정 영역 (2분할)
top_col1, top_col2 = st.columns([1, 1])

with top_col1:
    st.subheader("1. 원본 파일 업로드")
    uploaded_file = st.file_uploader("변환할 엑셀 파일을 올려주세요", type=['xlsx', 'xls'])

with top_col2:
    st.subheader("2. 정보 입력 및 변환 모드 선택")
    
    # 추가된 부분: 고객사명 및 제품명 입력
    customer_name = st.text_input("고객사명")
    product_name = st.text_input("제품명")
    
    # CFF와 HP를 선택할 수 있는 셀렉트박스
    mode = st.selectbox("업체 타입을 선택하세요", ["CFF", "HP"])
    
    # 선택된 모드에 따라 사용할 템플릿 파일명 미리 지정
    if mode == "CFF":
        st.info("💡 [CFF 모드] '83 CFF' 및 '26 통합' 양식으로 변환합니다.")
    else:
        st.info("💡 [HP 모드] '83 HP' 및 '26 통합' 양식으로 변환합니다.")

st.markdown("---")

# [하단] 실행 및 결과 영역 (2분할)
btm_col1, btm_col2 = st.columns([1, 1])

# 결과물을 담을 변수 초기화 (세션 스테이트 사용)
if 'result_83' not in st.session_state:
    st.session_state.result_83 = None
if 'result_26' not in st.session_state:
    st.session_state.result_26 = None
if 'fname_83' not in st.session_state:
    st.session_state.fname_83 = "83_Converted.xlsx"
if 'fname_26' not in st.session_state:
    st.session_state.fname_26 = "26_Converted.xlsx"

with btm_col1:
    st.subheader("3. 변환 실행")
    if st.button("변환 시작", type="primary", use_container_width=True):
        if uploaded_file is not None:
            try:
                # 원본 읽기
                input_df = pd.read_excel(uploaded_file)
                
                # 템플릿 경로 설정 (상대 경로)
                base_path = "template"
                
                if mode == "CFF":
                    # CFF 로직 실행
                    res_83 = logic_cff_83(input_df, os.path.join(base_path, "83 CFF.xlsx"), customer_name, product_name)
                    res_26 = logic_cff_26(input_df, os.path.join(base_path, "26 통합.xlsx"), customer_name, product_name)
                    
                    # CFF 파일명 지정
                    st.session_state.fname_83 = f"83 ALLERGENS {product_name}.xlsx"
                    st.session_state.fname_26 = f"ALLERGEN {product_name}.xlsx"
                else:
                    # HP 로직 실행
                    res_83 = logic_hp_83(input_df, os.path.join(base_path, "83 HP.xlsx"), customer_name, product_name)
                    res_26 = logic_hp_26(input_df, os.path.join(base_path, "26 통합.xlsx"), customer_name, product_name)
                    
                    # HP 임시 파일명
                    st.session_state.fname_83 = f"HP_83_Converted.xlsx"
                    st.session_state.fname_26 = f"HP_26_Converted.xlsx"
                
                # 결과를 세션에 저장 (화면이 리로딩돼도 다운로드 버튼 유지)
                st.session_state.result_83 = to_excel(res_83)
                st.session_state.result_26 = to_excel(res_26)
                
                st.success("변환이 완료되었습니다! 오른쪽에서 다운로드하세요. 👉")
                
            except Exception as e:
                st.error(f"오류가 발생했습니다: {e}")
        else:
            st.warning("먼저 원본 파일을 업로드해주세요.")

with btm_col2:
    st.subheader("4. 결과물 다운로드")
    
    if st.session_state.result_83 and st.session_state.result_26:
        prefix = "CFF" if mode == "CFF" else "HP"
        
        # 다운로드 버튼 1: 83 양식
        st.download_button(
            label=f"📥 {prefix}_83 양식 다운로드",
            data=st.session_state.result_83,
            file_name=st.session_state.fname_83,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
        # 다운로드 버튼 2: 26 통합 양식
        st.download_button(
            label=f"📥 {prefix}_26 통합 다운로드",
            data=st.session_state.result_26,
            file_name=st.session_state.fname_26,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    else:
        st.write("왼쪽에서 '변환 시작' 버튼을 누르면 다운로드 버튼이 나타납니다.")
